from __future__ import annotations

import csv
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import bad_request
from app.modules.database.engines import ExternalDatabase, placeholder_name, split_sql_statements
from app.modules.database.schemas import clean_identifier

BATCH_SIZE = 500


def import_data(runtime: ExternalDatabase, params: dict[str, Any], input_path: Path) -> int:
    fmt = str(params.get("format") or "").lower()
    database = str(params.get("database") or runtime.database or "")
    target_table = str(params.get("target_table") or "")
    mode = str(params.get("mode") or "append")
    mapping = params.get("mapping") or {}
    create_table = bool(params.get("create_table"))
    if fmt == "csv":
        return _import_csv(runtime, input_path, database, target_table, mode, mapping, create_table)
    if fmt == "xlsx":
        return _import_xlsx(runtime, input_path, database, target_table, mode, mapping, create_table)
    if fmt == "sql":
        return _import_sql(runtime, input_path)
    if fmt == "sqlite":
        return _import_sqlite(runtime, input_path, database, target_table, mode, mapping, create_table)
    raise bad_request("error.databaseImportFormatUnsupported", "Unsupported import format")


def _import_csv(
    runtime: ExternalDatabase,
    input_path: Path,
    database: str,
    target_table: str,
    mode: str,
    mapping: dict[str, str],
    create_table: bool,
) -> int:
    if not target_table:
        raise bad_request("error.databaseTargetTableRequired", "Target table is required")
    with input_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        target_columns = _target_columns(headers, mapping)
        _prepare_table(runtime, database, target_table, target_columns, mode, create_table)

    # Fast path: LOAD DATA LOCAL streams the whole file server-side (10-50x faster than
    # row INSERTs). Only for plain append/overwrite on MySQL/MariaDB with local_infile ON;
    # upsert keeps the row-by-row path because LOAD DATA has no ON DUPLICATE KEY UPDATE.
    if mode in {"append", "overwrite"} and runtime.supports_load_data():
        normalized = _normalize_newlines(input_path)
        try:
            return runtime.load_data_local(str(normalized), target_table, list(target_columns.values()), database)
        finally:
            if normalized != input_path:
                normalized.unlink(missing_ok=True)

    # Fallback: reuse one connection for the whole import instead of reconnecting per batch.
    with input_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        count = 0
        batch: list[dict[str, Any]] = []
        with runtime.session_connection() as conn:
            for row in reader:
                batch.append({target_columns[source]: value for source, value in row.items() if source in target_columns})
                if len(batch) >= BATCH_SIZE:
                    count += _insert_batch(runtime, database, target_table, batch, mode, conn)
                    batch = []
            count += _insert_batch(runtime, database, target_table, batch, mode, conn)
        return count


def _import_xlsx(
    runtime: ExternalDatabase,
    input_path: Path,
    database: str,
    target_table: str,
    mode: str,
    mapping: dict[str, str],
    create_table: bool,
) -> int:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    count = 0
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, [])]
            table = target_table or clean_identifier(sheet.title, "table")
            target_columns = _target_columns(headers, mapping)
            _prepare_table(runtime, database, table, target_columns, mode, create_table)
            batch: list[dict[str, Any]] = []
            for values in rows:
                record = {target_columns[headers[index]]: values[index] for index in range(min(len(headers), len(values))) if headers[index] in target_columns}
                batch.append(record)
                if len(batch) >= BATCH_SIZE:
                    count += _insert_batch(runtime, database, table, batch, mode)
                    batch = []
            count += _insert_batch(runtime, database, table, batch, mode)
    finally:
        workbook.close()
    return count


def _import_sql(runtime: ExternalDatabase, input_path: Path) -> int:
    content = input_path.read_text(encoding="utf-8")
    count = 0
    for statement in split_sql_statements(content):
        count += runtime.execute_write(statement)
    return count


def _import_sqlite(
    runtime: ExternalDatabase,
    input_path: Path,
    database: str,
    target_table: str,
    mode: str,
    mapping: dict[str, str],
    create_table: bool,
) -> int:
    count = 0
    with sqlite3.connect(input_path) as source:
        source.row_factory = sqlite3.Row
        tables = [target_table] if target_table else [row[0] for row in source.execute("SELECT name FROM sqlite_master WHERE type='table'")]
        for table in tables:
            src_table = clean_identifier(table, "table")
            rows = source.execute(f'SELECT * FROM "{src_table}"')
            headers = [column[0] for column in rows.description or []]
            target_columns = _target_columns(headers, mapping)
            _prepare_table(runtime, database, src_table, target_columns, mode, create_table)
            batch: list[dict[str, Any]] = []
            for row in rows:
                batch.append({target_columns[key]: row[key] for key in headers if key in target_columns})
                if len(batch) >= BATCH_SIZE:
                    count += _insert_batch(runtime, database, src_table, batch, mode)
                    batch = []
            count += _insert_batch(runtime, database, src_table, batch, mode)
    return count


def _target_columns(headers: list[str], mapping: dict[str, str]) -> dict[str, str]:
    result = {}
    for header in headers:
        source = str(header or "").strip()
        if not source:
            continue
        target = str(mapping.get(source) or source).strip()
        result[source] = clean_identifier(target, "column")
    if not result:
        raise bad_request("error.databaseImportNoColumns", "No import columns found")
    return result


def _prepare_table(
    runtime: ExternalDatabase,
    database: str,
    table: str,
    target_columns: dict[str, str],
    mode: str,
    create_table: bool,
) -> None:
    table = clean_identifier(table, "table")
    if create_table:
        columns_sql = ", ".join(f"{runtime.quote_identifier(column)} TEXT" for column in target_columns.values())
        runtime.execute_write(f"CREATE TABLE IF NOT EXISTS {runtime.qualified_table(table, database)} ({columns_sql})")
    if mode == "overwrite":
        if runtime.connection.db_type == "sqlite":
            runtime.execute_write(f"DELETE FROM {runtime.qualified_table(table, database)}")
        else:
            runtime.execute_write(f"TRUNCATE TABLE {runtime.qualified_table(table, database)}")


def _insert_batch(runtime: ExternalDatabase, database: str, table: str, rows: list[dict[str, Any]], mode: str, conn: Any = None) -> int:
    if not rows:
        return 0
    columns = list(rows[0].keys())
    table_sql = runtime.qualified_table(table, database)
    cols_sql = ", ".join(runtime.quote_identifier(column) for column in columns)
    placeholders = {column: placeholder_name(column) for column in columns}
    values_sql = ", ".join(f":{placeholders[column]}" for column in columns)
    prefix = "INSERT"
    suffix = ""
    if mode == "upsert":
        primary_keys = runtime.primary_keys(table, database)
        if runtime.connection.db_type == "sqlite":
            prefix = "INSERT OR REPLACE"
        elif primary_keys:
            update_columns = [column for column in columns if column not in primary_keys]
            suffix = " ON DUPLICATE KEY UPDATE " + ", ".join(
                f"{runtime.quote_identifier(column)} = VALUES({runtime.quote_identifier(column)})" for column in update_columns
            )
    sql = f"{prefix} INTO {table_sql} ({cols_sql}) VALUES ({values_sql}){suffix}"
    bound_rows = [{placeholders[column]: row.get(column) for column in columns} for row in rows]
    if conn is not None:
        return runtime.execute_many_on(conn, sql, bound_rows)
    return runtime.execute_many(sql, bound_rows)


def _normalize_newlines(input_path: Path) -> Path:
    """LOAD DATA uses LINES TERMINATED BY '\\n'; rewrite CRLF/CR files to LF (next to the
    original) so the last column never keeps a trailing '\\r'. Returns the original path
    when it is already LF-only."""
    raw = input_path.read_bytes()
    if b"\r" not in raw:
        return input_path
    fixed = raw.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
    target = input_path.with_suffix(input_path.suffix + ".lf")
    target.write_bytes(fixed)
    return target
