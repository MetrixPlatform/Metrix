from __future__ import annotations

import csv
import sqlite3
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import text

from app.core.exceptions import bad_request
from app.modules.database.engines import ExternalDatabase, classify_sql, rows_to_dicts
from app.modules.database.schemas import clean_identifier

BATCH_SIZE = 1000


def export_data(runtime: ExternalDatabase, params: dict[str, Any], output_path: Path) -> int:
    fmt = str(params.get("format") or "").lower()
    database = str(params.get("database") or runtime.database or "")
    tables = [clean_identifier(str(table), "table") for table in params.get("tables") or []]
    sql = str(params.get("sql") or "").strip()
    queries = _normalize_queries(params.get("queries") or [])
    if fmt == "csv":
        return _export_csv(runtime, output_path, database, tables, sql, queries)
    if fmt == "xlsx":
        return _export_xlsx(runtime, output_path, database, tables, sql, queries)
    if fmt == "sqlite":
        return _export_sqlite(runtime, output_path, database, tables, sql, queries)
    if fmt == "sql":
        return _export_sql(runtime, output_path, database, tables, sql, queries)
    raise bad_request("error.databaseExportFormatUnsupported", "Unsupported export format")


def _export_csv(runtime: ExternalDatabase, output_path: Path, database: str, tables: list[str], sql: str, queries: list[tuple[str, str]]) -> int:
    datasets = _datasets(runtime, database, tables, sql, queries)
    if len(datasets) != 1:
        raise bad_request("error.databaseCsvSingleOnly", "CSV export supports one table or one query")
    row_count = 0
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer: csv.DictWriter | None = None
        for _, _, rows in datasets:
            for batch in rows:
                if not batch:
                    continue
                if writer is None:
                    writer = csv.DictWriter(file, fieldnames=list(batch[0].keys()))
                    writer.writeheader()
                writer.writerows(batch)
                row_count += len(batch)
    return row_count


def _export_xlsx(runtime: ExternalDatabase, output_path: Path, database: str, tables: list[str], sql: str, queries: list[tuple[str, str]]) -> int:
    workbook = Workbook(write_only=True)
    row_count = 0
    for name, columns, rows in _datasets(runtime, database, tables, sql, queries):
        sheet = workbook.create_sheet(_safe_sheet_name(name))
        sheet.append(columns)
        for batch in rows:
            for row in batch:
                sheet.append([row.get(column) for column in columns])
            row_count += len(batch)
    if not workbook.sheetnames:
        workbook.create_sheet("empty")
    workbook.save(output_path)
    return row_count


def _export_sqlite(runtime: ExternalDatabase, output_path: Path, database: str, tables: list[str], sql: str, queries: list[tuple[str, str]]) -> int:
    row_count = 0
    with sqlite3.connect(output_path) as target:
        for name, columns, rows in _datasets(runtime, database, tables, sql, queries):
            table_name = _safe_sqlite_identifier(name)
            target.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            target.execute(f'CREATE TABLE "{table_name}" ({", ".join(f"{_q_sqlite(column)} TEXT" for column in columns)})')
            placeholders = ", ".join("?" for _ in columns)
            insert_sql = f'INSERT INTO "{table_name}" ({", ".join(_q_sqlite(column) for column in columns)}) VALUES ({placeholders})'
            for batch in rows:
                target.executemany(insert_sql, [[row.get(column) for column in columns] for row in batch])
                row_count += len(batch)
        target.commit()
    return row_count


def _export_sql(runtime: ExternalDatabase, output_path: Path, database: str, tables: list[str], sql: str, queries: list[tuple[str, str]]) -> int:
    row_count = 0
    with output_path.open("w", encoding="utf-8", newline="\n") as file:
        for name, columns, rows in _datasets(runtime, database, tables, sql, queries):
            table_name = runtime.quote_identifier(_safe_sql_identifier(name))
            file.write(f"DROP TABLE IF EXISTS {table_name};\n")
            file.write(f"CREATE TABLE {table_name} ({', '.join(runtime.quote_identifier(column) + ' TEXT' for column in columns)});\n")
            for batch in rows:
                for row in batch:
                    values = ", ".join(_sql_literal(row.get(column)) for column in columns)
                    file.write(f"INSERT INTO {table_name} ({', '.join(runtime.quote_identifier(column) for column in columns)}) VALUES ({values});\n")
                row_count += len(batch)
    return row_count


def _normalize_queries(raw: list[Any]) -> list[tuple[str, str]]:
    queries: list[tuple[str, str]] = []
    for index, item in enumerate(raw, start=1):
        statement = str(item.get("sql") if isinstance(item, dict) else "").strip()
        if not statement:
            continue
        name = str(item.get("name") or "").strip() if isinstance(item, dict) else ""
        queries.append((name or f"query_{index}", statement))
    return queries


def _datasets(
    runtime: ExternalDatabase,
    database: str,
    tables: list[str],
    sql: str,
    queries: list[tuple[str, str]],
) -> list[tuple[str, list[str], Iterable[list[dict[str, Any]]]]]:
    if queries:
        datasets = []
        for name, statement in queries:
            if classify_sql(statement) != "read":
                raise bad_request("error.databaseExportReadOnly", "Only read SQL can be exported")
            datasets.append((name, _query_columns(runtime, statement), _query_batches(runtime, statement)))
        return datasets
    if sql:
        if classify_sql(sql) != "read":
            raise bad_request("error.databaseExportReadOnly", "Only read SQL can be exported")
        return [("query", _query_columns(runtime, sql), _query_batches(runtime, sql))]
    selected_tables = tables or runtime.tables(database)
    return [
        (table, [column.name for column in runtime.columns(table, database)], _table_batches(runtime, database, table))
        for table in selected_tables
    ]


def _query_columns(runtime: ExternalDatabase, sql: str) -> list[str]:
    with runtime.engine.connect() as conn:
        result = conn.execute(text(f"SELECT * FROM ({sql.strip().rstrip(';')}) AS mtx_export LIMIT 0"))
        return list(result.keys())


def _query_batches(runtime: ExternalDatabase, sql: str):
    statement = sql.strip().rstrip(";")
    with runtime.engine.connect() as conn:
        result = conn.execution_options(stream_results=True).execute(text(statement)).mappings()
        while True:
            rows = result.fetchmany(BATCH_SIZE)
            if not rows:
                break
            yield rows_to_dicts(rows)


def _table_batches(runtime: ExternalDatabase, database: str, table: str):
    statement = f"SELECT * FROM {runtime.qualified_table(table, database)}"
    with runtime.engine.connect() as conn:
        result = conn.execution_options(stream_results=True).execute(text(statement)).mappings()
        while True:
            rows = result.fetchmany(BATCH_SIZE)
            if not rows:
                break
            yield rows_to_dicts(rows)


def _safe_sheet_name(value: str) -> str:
    cleaned = "".join("_" if char in r'[]:*?/\\' else char for char in value)[:31]
    return cleaned or "sheet"


def _safe_sqlite_identifier(value: str) -> str:
    return "".join(char if char.isalnum() or char == "_" else "_" for char in value) or "data"


def _safe_sql_identifier(value: str) -> str:
    return "".join(char if char.isalnum() or char == "_" else "_" for char in value) or "data"


def _q_sqlite(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"
