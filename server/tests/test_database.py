import time
from datetime import datetime
from pathlib import Path

from sqlalchemy import create_engine, text

from app.core.security import decrypt_secret, encrypt_secret
from app.modules.database.engines import ExternalDatabase
from app.modules.database.importers import _insert_batch
from app.modules.database.models import DatabaseConnection
from test_auth_rbac import create_client, install_sqlite, login


def make_target_db(path: Path) -> None:
    engine = create_engine(f"sqlite:///{path}")
    with engine.begin() as conn:
        conn.execute(text("CREATE TABLE people (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, age INTEGER)"))
        conn.execute(text("INSERT INTO people (name, age) VALUES ('Alice', 20), ('Bob', 30)"))
    engine.dispose()


def create_sqlite_connection(conn_id: str, target_path: Path) -> int:
    from app.core.install import load_install_config

    engine = create_engine(load_install_config().database_url)
    encrypted = encrypt_secret("unused")
    with engine.begin() as conn:
        conn.execute(
            text(
                "INSERT INTO db_connections "
                "(conn_id, name, db_type, host, port, username, password_encrypted, default_database, is_shared, is_active, created_by, created_at, updated_at) "
                "VALUES (:conn_id, :name, 'sqlite', :host, 0, 'sqlite', :password, '', 1, 1, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
            ),
            {"conn_id": conn_id, "name": "SQLite Target", "host": str(target_path), "password": encrypted},
        )
        connection_id = conn.execute(text("SELECT id FROM db_connections WHERE conn_id = :conn_id"), {"conn_id": conn_id}).scalar_one()
    engine.dispose()
    assert decrypt_secret(encrypted) == "unused"
    return connection_id


def test_database_metadata_query_rows_and_export(tmp_path, monkeypatch):
    client = create_client(tmp_path, monkeypatch)
    payload = install_sqlite(client, tmp_path)
    headers = login(client, payload["admin_username"], payload["admin_password"])
    target_path = tmp_path / "target.db"
    make_target_db(target_path)
    connection_id = create_sqlite_connection("db_sqlite_test", target_path)
    settings_payload = client.get("/api/settings", headers=headers).json()
    settings_payload["data_job_retention_hours"] = 30 * 24
    assert client.put("/api/settings", json=settings_payload, headers=headers).status_code == 200

    schemas = client.get("/api/databases/db_sqlite_test/schemas", headers=headers)
    assert schemas.status_code == 200
    assert schemas.json() == [{"name": "main"}]

    tables = client.get("/api/databases/db_sqlite_test/tables", headers=headers)
    assert tables.status_code == 200
    assert {"name": "people"} in tables.json()

    numeric_schema_tables = client.get("/api/databases/db_sqlite_test/tables?database=0421", headers=headers)
    assert numeric_schema_tables.status_code == 200

    invalid_schema_tables = client.get("/api/databases/db_sqlite_test/tables?database=bad-name", headers=headers)
    assert invalid_schema_tables.status_code == 400

    table_data = client.get("/api/databases/db_sqlite_test/table-data?table=people&page_size=10", headers=headers)
    assert table_data.status_code == 200
    assert table_data.json()["total"] == 2
    assert table_data.json()["total_exact"] is True
    assert table_data.json()["primary_keys"] == ["id"]

    fast_table_data = client.get("/api/databases/db_sqlite_test/table-data?table=people&page_size=1&include_total=false", headers=headers)
    assert fast_table_data.status_code == 200
    assert fast_table_data.json()["total"] == 2
    assert fast_table_data.json()["total_exact"] is False

    ordered_desc = client.get("/api/databases/db_sqlite_test/table-data?table=people&order_by=name&order_desc=true&page_size=10", headers=headers)
    assert ordered_desc.status_code == 200
    assert [row["name"] for row in ordered_desc.json()["rows"]] == ["Bob", "Alice"]
    ordered_asc = client.get("/api/databases/db_sqlite_test/table-data?table=people&order_by=name&page_size=10", headers=headers)
    assert [row["name"] for row in ordered_asc.json()["rows"]] == ["Alice", "Bob"]

    selected = client.post("/api/databases/db_sqlite_test/query", json={"sql": "SELECT name FROM people ORDER BY id"}, headers=headers)
    assert selected.status_code == 200
    assert [row["name"] for row in selected.json()["rows"]] == ["Alice", "Bob"]

    created = client.post(
        "/api/databases/db_sqlite_test/table-rows",
        json={"table": "people", "values": {"name": "Cara", "age": 25}},
        headers=headers,
    )
    assert created.status_code == 200

    updated = client.put(
        "/api/databases/db_sqlite_test/table-rows",
        json={"table": "people", "keys": {"id": 3}, "values": {"name": "Cara", "age": 26}},
        headers=headers,
    )
    assert updated.status_code == 200

    script = client.post(
        "/api/databases/db_sqlite_test/run-script",
        json={"content": "SELECT COUNT(*) AS total FROM people; UPDATE people SET age = age + 1 WHERE id = 3;"},
        headers=headers,
    )
    assert script.status_code == 200
    assert len(script.json()["results"]) == 2

    submitted = client.post(
        "/api/databases/db_sqlite_test/export",
        json={"format": "csv", "tables": ["people"]},
        headers=headers,
    )
    assert submitted.status_code == 200
    job_id = submitted.json()["job_id"]
    job = wait_job(client, headers, job_id)
    assert job["status"] == "success"
    assert job["row_count"] == 3
    retention_seconds = (datetime.fromisoformat(job["expires_at"]) - datetime.fromisoformat(job["finished_at"])).total_seconds()
    assert 29 * 24 * 60 * 60 < retention_seconds <= 30 * 24 * 60 * 60

    submitted_later = client.post(
        "/api/databases/db_sqlite_test/export",
        json={"format": "csv", "tables": ["people"]},
        headers=headers,
    )
    assert submitted_later.status_code == 200
    later_job_id = submitted_later.json()["job_id"]
    later_job = wait_job(client, headers, later_job_id)
    assert later_job["status"] == "success"

    jobs_asc = client.get("/api/database-transfer-jobs?sort_order=ascend", headers=headers)
    assert jobs_asc.status_code == 200
    assert [item["job_id"] for item in jobs_asc.json()["items"][:2]] == [job_id, later_job_id]

    jobs_desc = client.get("/api/database-transfer-jobs?sort_order=descend", headers=headers)
    assert jobs_desc.status_code == 200
    assert [item["job_id"] for item in jobs_desc.json()["items"][:2]] == [later_job_id, job_id]

    download_count = client.get("/api/database-transfer-jobs/download-count", headers=headers)
    assert download_count.status_code == 200
    assert download_count.json()["count"] == 2
    scoped_count = client.get(f"/api/database-transfer-jobs/download-count?connection_id={connection_id}", headers=headers)
    assert scoped_count.status_code == 200
    assert scoped_count.json()["count"] == 2

    from app.core.install import load_install_config

    engine = create_engine(load_install_config().database_url)
    with engine.begin() as conn:
        conn.execute(
            text(
                "INSERT INTO users "
                "(username, full_name, phone, email, company, department, password_hash, approval_status, is_active, is_builtin, rejected_reason, created_at, updated_at) "
                "VALUES ('other_owner', 'Other Owner', '', '', '', '', 'hash', 'approved', 1, 0, '', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
            )
        )
        other_user_id = conn.execute(text("SELECT id FROM users WHERE username = 'other_owner'")).scalar_one()
        conn.execute(
            text(
                "INSERT INTO data_jobs "
                "(job_id, kind, connection_id, format, params_json, status, file_name, file_path, file_size, row_count, error_code, created_by, created_at, finished_at, expires_at) "
                "VALUES ('other-export-job', 'export', :connection_id, 'csv', '{}', 'success', 'other-export.csv', '', 0, 0, '', :created_by, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
            ),
            {"connection_id": connection_id, "created_by": other_user_id},
        )
    engine.dispose()

    own_jobs = client.get("/api/database-transfer-jobs?created_by=me", headers=headers)
    assert own_jobs.status_code == 200
    assert all(item["created_by_username"] == payload["admin_username"] for item in own_jobs.json()["items"])
    other_jobs = client.get("/api/database-transfer-jobs?created_by=others", headers=headers)
    assert other_jobs.status_code == 200
    assert [item["job_id"] for item in other_jobs.json()["items"]] == ["other-export-job"]
    assert other_jobs.json()["items"][0]["created_by_username"] == "other_owner"
    keyword_jobs = client.get("/api/database-transfer-jobs?keyword=other-export", headers=headers)
    assert keyword_jobs.status_code == 200
    assert [item["job_id"] for item in keyword_jobs.json()["items"]] == ["other-export-job"]
    scoped_jobs = client.get(f"/api/database-transfer-jobs?connection_id={connection_id}", headers=headers)
    assert scoped_jobs.status_code == 200
    assert scoped_jobs.json()["total"] >= 3

    download = client.get(f"/api/database-transfer-jobs/{job_id}/download", headers=headers)
    assert download.status_code == 200
    assert b"Alice" in download.content
    assert client.get("/api/database-transfer-jobs/download-count", headers=headers).json()["count"] == 1


def wait_job(client, headers, job_id: str):
    for _ in range(30):
        response = client.get(f"/api/database-transfer-jobs/{job_id}", headers=headers)
        assert response.status_code == 200
        payload = response.json()
        if payload["status"] in {"success", "failed"}:
            return payload
        time.sleep(0.1)
    raise AssertionError("data job did not finish")


def test_database_export_multiple_queries(tmp_path, monkeypatch):
    import io

    from openpyxl import load_workbook

    client = create_client(tmp_path, monkeypatch)
    payload = install_sqlite(client, tmp_path)
    headers = login(client, payload["admin_username"], payload["admin_password"])
    target_path = tmp_path / "multi.db"
    make_target_db(target_path)
    create_sqlite_connection("db_multi_test", target_path)

    submitted = client.post(
        "/api/databases/db_multi_test/export",
        json={
            "format": "xlsx",
            "queries": [
                {"name": "names", "sql": "SELECT name FROM people"},
                {"name": "ages", "sql": "SELECT age FROM people"},
            ],
        },
        headers=headers,
    )
    assert submitted.status_code == 200
    job = wait_job(client, headers, submitted.json()["job_id"])
    assert job["status"] == "success"
    assert job["row_count"] == 4

    download = client.get(f"/api/database-transfer-jobs/{job['job_id']}/download", headers=headers)
    assert download.status_code == 200
    workbook = load_workbook(io.BytesIO(download.content), read_only=True)
    assert set(workbook.sheetnames) == {"names", "ages"}
    workbook.close()

    csv_multi = client.post(
        "/api/databases/db_multi_test/export",
        json={
            "format": "csv",
            "queries": [
                {"name": "names", "sql": "SELECT name FROM people"},
                {"name": "ages", "sql": "SELECT age FROM people"},
            ],
        },
        headers=headers,
    )
    assert csv_multi.status_code == 200
    csv_job = wait_job(client, headers, csv_multi.json()["job_id"])
    assert csv_job["status"] == "failed"


def test_sql_script_is_database_scoped(tmp_path, monkeypatch):
    client = create_client(tmp_path, monkeypatch)
    payload = install_sqlite(client, tmp_path)
    headers = login(client, payload["admin_username"], payload["admin_password"])
    connection_id = create_sqlite_connection("db_scripts_test", tmp_path / "scripts.db")

    first = client.post(
        "/api/sql-scripts",
        json={"name": "report_0421", "content": "SELECT 1", "connection_id": connection_id, "database": "0421"},
        headers=headers,
    )
    assert first.status_code == 200
    assert first.json()["database"] == "0421"
    second = client.post(
        "/api/sql-scripts",
        json={"name": "report_other", "content": "SELECT 2", "connection_id": connection_id, "database": "report"},
        headers=headers,
    )
    assert second.status_code == 200

    scoped = client.get(f"/api/sql-scripts?connection_id={connection_id}&database=0421", headers=headers)
    assert scoped.status_code == 200
    assert [item["name"] for item in scoped.json()["items"]] == ["report_0421"]

    all_scripts = client.get(f"/api/sql-scripts?connection_id={connection_id}", headers=headers)
    assert {item["name"] for item in all_scripts.json()["items"]} == {"report_0421", "report_other"}


def test_database_import_uses_safe_bind_names(tmp_path):
    target_path = tmp_path / "bind_names.db"
    engine = create_engine(f"sqlite:///{target_path}")
    with engine.begin() as conn:
        conn.execute(text('CREATE TABLE special ("value$" TEXT)'))
    engine.dispose()

    connection = DatabaseConnection(
        conn_id="bind_names",
        name="Bind Names",
        db_type="sqlite",
        host=str(target_path),
        port=0,
        username="sqlite",
        password_encrypted="",
        default_database="",
        is_shared=True,
        is_active=True,
        created_by=1,
    )
    with ExternalDatabase(connection, "") as runtime:
        assert _insert_batch(runtime, "", "special", [{"value$": "ok"}], "append") == 1
        columns, rows, total, _ = runtime.execute_sql('SELECT "value$" AS value FROM special')

    assert columns == ["value"]
    assert rows == [{"value": "ok"}]
    assert total == 1
